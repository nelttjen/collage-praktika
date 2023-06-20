from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell, MergedCell
from typing import Union, List

wb: Workbook = load_workbook('wb.xlsx')
sheet: Worksheet = wb.get_sheet_by_name("Лист")
rows = list(sheet.iter_rows())[11:-1]
formatted_rows = []
for row in rows:
    row: List[Union[Cell, MergedCell]]
    for cell in row:
        if val := cell.value:
            items = val.split("(")
            items = [i.replace(")", '').strip() for i in items]
            name, total, process, total_registered, total_failed, total_ignored = items

            total = int(total)
            process = int(float(process.split(" = ")[1].replace(",", '.')))
            total_registered = int(float(total_registered.split(" = ")[1].replace(",", '.')))
            total_failed = int(float(total_failed.split(" = ")[1].replace(",", '.')))
            total_ignored = int(float(total_ignored.split(" = ")[1].replace(",", '.')))
            formatted_rows.append([name, total, process, total_registered, total_failed, total_ignored])

out_wb: Workbook = load_workbook('out_wb.xlsx')
out_sheet: Worksheet = out_wb.get_sheet_by_name('Отправка РЭМД август 2022')
out_rows = list(out_sheet.iter_rows())
names = [i[1].value for i in out_rows[3:]]
names_dict = {}
for i, name in enumerate(names):
    names_dict[name] = i + 4

last_index = 0

for i in range(1, 99999):
    cell = out_sheet.cell(1, i)
    if cell.value is None and not type(cell) == MergedCell:
        last_index = i
        break

last_item_index = len(list(out_sheet.rows))
processed_names = []


for item in formatted_rows:
    name, total, process, total_registered, total_failed, total_ignored = item
    if name in names_dict.keys():
        item_index = names_dict[name]
        found = True
    else:
        last_item_index += 1
        item_index = last_item_index * 1
        found = False

    if not found:
        item_id = item_index - 3
        out_sheet.cell(item_index, 1, value=item_id)
        out_sheet.cell(item_index, 2, value=name)
        for i in range(4, last_index + 1):
            out_sheet.cell(item_index, i, value=0)

    out_sheet.cell(item_index, last_index, value=total)
    out_sheet.cell(item_index, last_index+1, value=process)
    out_sheet.cell(item_index, last_index+2, value=total_registered)
    out_sheet.cell(item_index, last_index+3, value=total_failed)
    out_sheet.cell(item_index, last_index+4, value=total_ignored)
    processed_names.append(name)

for name in names_dict.keys():
    if name not in processed_names:
        index = names_dict[name]
        out_sheet.cell(index, last_index, value=0)
        out_sheet.cell(index, last_index + 1, value=0)
        out_sheet.cell(index, last_index + 2, value=0)
        out_sheet.cell(index, last_index + 3, value=0)
        out_sheet.cell(index, last_index + 4, value=0)

cell_names = ['Общее количество', 'Всего в процессе регистрации', 'Всего зарегистрировано в РЭМД', 'Всего отказано в регистрации', 'Не отправлен на регистрацию']

result_name = input('Название итога: ')

for i in range(5):
    cell: Cell = out_sheet.cell(1, last_index+i)
    cell.value = f'=SUM({cell.column_letter}4:{cell.column_letter}{last_item_index})'

out_sheet.cell(2, last_index, value=result_name)
out_sheet.merge_cells(start_row=2, end_row=2, start_column=last_index, end_column=last_index + 4)

for i, cell_name in enumerate(cell_names):
    out_sheet.cell(3, last_index+i, value=cell_name)

out_wb.save('out_wb2.xlsx')